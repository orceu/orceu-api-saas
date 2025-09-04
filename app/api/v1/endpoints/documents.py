# app/api/v1/endpoints/documents.py

from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, status
from pydantic import BaseModel
from typing import Dict
from uuid import uuid4
from openpyxl import load_workbook
import os
import tempfile
import pandas as pd
import chardet
import csv
import math

from app.core.dependencies import get_tenant
from app.services.storage_manager import S3FileManager

router = APIRouter()
s3_manager = S3FileManager()


class SchemaInferRequest(BaseModel):
    sheet_name: str
    column_mapping: Dict[int, str]  # índice -> nome do campo
    required_field: str             # nome do campo que não pode ser nulo


def _sanitize_for_json(obj):
    """Converte NaN/Inf para None em listas, dicts ou floats isolados."""
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    if isinstance(obj, list):
        return [_sanitize_for_json(x) for x in obj]
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    return obj

def _read_excel_with_merges(path: str, max_rows: int = 20):
    """
    Lê Excel preservando merges (colspan) e retorna primeiras linhas de cada sheet.
    """
    wb = load_workbook(path, read_only=False, data_only=True)
    data_by_sheet = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Mapear merges: todas as células apontam para a top-left
        merged_map = {}
        for merged in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
            top_left = (min_row, min_col)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if (r, c) != top_left:
                        merged_map[(r, c)] = top_left

        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=False)):
            values = []
            for j, cell in enumerate(row, start=1):
                coord = (i + 1, j)
                if coord in merged_map:
                    master = merged_map[coord]
                    master_cell = ws.cell(*master)
                    values.append(master_cell.value)
                else:
                    values.append(cell.value)
            rows.append(values)

        data_by_sheet[sheet_name] = rows

    wb.close()
    return data_by_sheet

def read_excel_preview(file_path: str, max_rows: int = 20, sheet_index: int = 0):
    # Carrega workbook sem read_only para acessar merges
    wb = load_workbook(file_path, data_only=True)
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[sheet_index]]

    rows_data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        # Converte None -> "" igual ao PHP (ou "None" se preferir)
        row_values = [str(cell) if cell is not None else "" for cell in row]
        rows_data.append(row_values)

    result = {
        "sheet_name": sheetnames[sheet_index],
        "total_rows": ws.max_row,
        "sample_rows": rows_data
    }
    return result


# ============================================
# DOCUMENTS PIPELINE
# ============================================

@router.post("/documents/upload", status_code=status.HTTP_201_CREATED)
def upload_document(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant)
):
    """
    1) Gera UUID para o documento.
    2) Cria 'pasta' no S3 para o tenant/document_id (prefixo).
    3) Faz upload do arquivo (usando o document_id como nome).
    4) Retorna metadados + URL de download temporária.
    """
    document_id = str(uuid4())

    prefix = f"documents/{tenant_id}/"

    # cria arquivo temporário local
    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        tmp.write(file.file.read())
        tmp_path = tmp.name

    # usa document_id como nome final, preservando a extensão
    ext = os.path.splitext(file.filename)[1]
    object_name = f"{prefix}{document_id}{ext}"

    # upload para o S3
    s3_manager.upload_file(tmp_path, object_name)

    # opcional: cria marcador de pasta
    s3_manager.upload_file(tmp_path, f"{prefix}.keep")

    os.remove(tmp_path)

    download_url = s3_manager.generate_presigned_url(object_name)

    return {
        "document_id": document_id,
        "tenant_id": tenant_id,
        "object_name": object_name,
        "download_url": download_url,
        "message": "Documento recebido e armazenado com sucesso."
    }


@router.get("/documents/imports/{document_id}/preview")
def preview_document(
    document_id: str,
    tenant_id: str = Depends(get_tenant)
):
    """
    Detecta:
    - sheets (Excel)
    - amostra de 20 primeiras linhas (linha a linha, célula a célula)
    - delimitador (CSV)
    - encoding
    - total de linhas
    """
    import chardet, csv
    from openpyxl import load_workbook

    # tenta localizar o arquivo no S3
    possible_exts = [".xlsx", ".xls", ".csv"]
    found = None
    for ext in possible_exts:
        key = f"documents/{tenant_id}/{document_id}{ext}"
        files = s3_manager.list_folder(os.path.dirname(key))
        if any(f == key for f in files):
            found = key
            break
    if not found:
        raise HTTPException(404, detail="Documento não encontrado no storage.")

    object_name = found
    ext = os.path.splitext(object_name)[1].lower()
    tmp_path = os.path.join(tempfile.gettempdir(), f"{uuid4()}{ext}")

    try:
        # baixa o arquivo do S3
        s3_manager.download_file(object_name, tmp_path)

        # Caso CSV
        if ext == ".csv":
            with open(tmp_path, "rb") as f:
                raw = f.read(5000)
                enc_guess = chardet.detect(raw)
            encoding = enc_guess.get("encoding", "utf-8")

            with open(tmp_path, "r", encoding=encoding, errors="ignore") as f:
                sample = f.read(4096)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter if sample else ","

            rows = []
            with open(tmp_path, "r", encoding=encoding, errors="ignore") as f:
                reader = csv.reader(f, delimiter=delimiter)
                for i, row in enumerate(reader):
                    if i >= 20:
                        break
                    # Trim + substitui None por ""
                    rows.append([str(cell).strip() if cell is not None else "" for cell in row])

            total_rows = sum(1 for _ in open(tmp_path, encoding=encoding, errors="ignore"))

            return {
                "document_id": document_id,
                "type": "csv",
                "sample_rows": rows,
                "delimiter": delimiter,
                "encoding": encoding,
                "total_rows": total_rows
            }

        # Caso Excel
        elif ext in [".xlsx", ".xls"]:
            wb = load_workbook(tmp_path, data_only=True)
            result_sheets = {}

            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 20:
                        break
                    # Trim + None → ""
                    rows.append([str(cell).strip() if cell is not None else "" for cell in row])

                result_sheets[sheetname] = {
                    "sample_rows": rows,
                    "total_rows": ws.max_row
                }

            return {
                "document_id": document_id,
                "type": "excel",
                "sheets": result_sheets
            }

        else:
            raise HTTPException(400, detail="Extensão não suportada.")

    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except PermissionError:
                pass

@router.post("/documents/imports/{document_id}/schema:infer")
def infer_schema(
    document_id: str,
    payload: SchemaInferRequest,
    tenant_id: str = Depends(get_tenant)
):
    """
    Lê a sheet escolhida, aplica o mapeamento de índices para campos,
    remove linhas onde `required_field` é nulo e retorna toda a planilha.
    """
    # localizar arquivo Excel no S3
    possible_exts = [".xlsx", ".xls"]
    found = None
    for ext in possible_exts:
        key = f"documents/{tenant_id}/{document_id}{ext}"
        files = s3_manager.list_folder(os.path.dirname(key))
        if any(f == key for f in files):
            found = key
            break
    if not found:
        raise HTTPException(404, detail="Documento não encontrado no storage.")

    tmp_path = os.path.join(tempfile.gettempdir(), f"{uuid4()}{ext}")
    try:
        s3_manager.download_file(found, tmp_path)

        wb = load_workbook(tmp_path, data_only=True)
        if payload.sheet_name not in wb.sheetnames:
            raise HTTPException(400, detail=f"Sheet '{payload.sheet_name}' não encontrada.")

        ws = wb[payload.sheet_name]

        mapped_rows = []
        for row in ws.iter_rows(values_only=True):
            mapped_row = {}
            for col_idx, field_name in payload.column_mapping.items():
                col_idx = int(col_idx)
                val = row[col_idx] if col_idx < len(row) else None
                if isinstance(val, str):
                    val = val.strip()
                mapped_row[field_name] = val

            # descarta linhas se o campo obrigatório estiver vazio/nulo
            if not mapped_row.get(payload.required_field):
                continue

            mapped_rows.append(mapped_row)

        return {
            "document_id": document_id,
            "sheet_name": payload.sheet_name,
            "mapping": payload.column_mapping,
            "required_field": payload.required_field,
            "rows": mapped_rows
        }

    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except PermissionError:
                pass


@router.post("/documents/imports/{document_id}/mapping")
def update_mapping(document_id: str, mapping: dict, tenant_id: str = Depends(get_tenant)):
    """
    Recebe ou atualiza mapeamento coluna → campo do target + transformações.
    """
    # TODO: salvar mapping em banco
    return {"document_id": document_id, "mapping": mapping, "message": "Mapping atualizado."}


@router.post("/documents/imports/{document_id}/validate")
def validate_document(document_id: str, tenant_id: str = Depends(get_tenant)):
    """
    Valida o arquivo (ou lote) contra o schema do target.
    """
    # TODO: rodar validação real
    return {
        "document_id": document_id,
        "valid": True,
        "errors": [],
        "warnings": []
    }


@router.post("/documents/imports/{document_id}/simulate")
def simulate_import(document_id: str, tenant_id: str = Depends(get_tenant)):
    """
    Simula a importação:
    - contabiliza creates/updates/dups/erros sem gravar.
    """
    # TODO: lógica real de simulação
    return {
        "document_id": document_id,
        "creates": 10,
        "updates": 2,
        "duplicates": 1,
        "errors": 0
    }


@router.post("/documents/imports/{document_id}/process")
def process_import(document_id: str, tenant_id: str = Depends(get_tenant)):
    """
    Enfileira a importação real.
    """
    run_id = str(uuid4())
    # TODO: enfileirar no worker
    return {"document_id": document_id, "run_id": run_id, "status": "queued"}


@router.get("/documents/imports/{document_id}/runs/{run_id}")
def get_import_run_status(document_id: str, run_id: str, tenant_id: str = Depends(get_tenant)):
    """
    Retorna status, métricas e erros exportáveis de uma execução de importação.
    """
    # TODO: consultar status real do run
    return {
        "document_id": document_id,
        "run_id": run_id,
        "status": "running",
        "metrics": {"processed": 100, "errors": 2},
        "errors": [
            {"row": 10, "message": "Código inválido"},
            {"row": 25, "message": "Unidade não reconhecida"}
        ]
    }
